from openpyxl import load_workbook
import json
import sys
import os

def populate_wbs_template(wbs_elements, template_path, output_path):
    """
    Populate the WBS template with WBS Elements data.
    
    Args:
        wbs_elements: List of dictionaries containing WBS Element data
        template_path: Path to the template file
        output_path: Path where the populated file should be saved
    """
    try:
        # Load the template
        wb = load_workbook(template_path, keep_vba=True)  # Keep VBA/Macro content
        
        # Select the Shared Template sheet
        ws = wb['Shared Template']
        
        # Start populating from row 18
        start_row = 18
        
        # Populate data
        for idx, element in enumerate(wbs_elements, start=start_row):
            # Map the data to the correct columns based on the template structure
            # Column B: Kind of change to process (Type)
            ws.cell(row=idx, column=2, value=element.get('type', ''))
            
            # Column C: System (Standard)
            ws.cell(row=idx, column=3, value='Standard')
            
            # Column D: Controlling Area
            ws.cell(row=idx, column=4, value=element.get('controllingArea', ''))
            
            # Column E: Company Code
            ws.cell(row=idx, column=5, value=element.get('companyCode', ''))
            
            # Column F: Project Name
            ws.cell(row=idx, column=6, value=element.get('projectName', ''))
            
            # Column G: Project Definition/Project Element
            ws.cell(row=idx, column=7, value=element.get('projectDefinition', ''))
            
            # Column H: Level
            ws.cell(row=idx, column=8, value=element.get('level', ''))
            
            # Column I: Project Type
            ws.cell(row=idx, column=9, value=element.get('projectType', ''))
            
            # Column J: Investment profile
            ws.cell(row=idx, column=10, value='')  # Left empty as per template
            
            # Column K: Responsible Profit Center (Responsible PC/CC)
            ws.cell(row=idx, column=11, value=element.get('responsiblePCCC', ''))
            
            # Column L: Planning element
            ws.cell(row=idx, column=12, value='X' if element.get('planningElement') else '')
            
            # Column M: Rubric element
            ws.cell(row=idx, column=13, value='X' if element.get('rubricElement') else '')
            
            # Column N: Billing element
            ws.cell(row=idx, column=14, value='X' if element.get('billingElement') else '')
            
            # Column O: Settlement Rule %
            ws.cell(row=idx, column=15, value=element.get('settlementRulePercent', ''))
            
            # Column P: Settlement Rule Goal
            ws.cell(row=idx, column=16, value=element.get('settlementRuleGoal', ''))
            
            # Column Q: Responsible Person
            ws.cell(row=idx, column=17, value=element.get('responsiblePerson', ''))
            
            # Column R: User ID
            ws.cell(row=idx, column=18, value=element.get('userId', ''))
            
            # Column S: Employment Number
            ws.cell(row=idx, column=19, value=element.get('employmentNumber', ''))
            
            # Column T: Functional Area
            ws.cell(row=idx, column=20, value=element.get('functionalArea', ''))
            
            # Column U: Project Specification
            ws.cell(row=idx, column=21, value=element.get('projectSpec', ''))
            
            # Column V: Mother Code
            ws.cell(row=idx, column=22, value=element.get('motherCode', ''))
            
            # Column W: TG Phase
            ws.cell(row=idx, column=23, value=element.get('tgPhase', ''))
            
            # Column X: Comment
            ws.cell(row=idx, column=24, value=element.get('comment', ''))
        
        # Save the workbook to the output path
        wb.save(output_path)
        print("Template has been populated successfully!")
        return True
        
    except Exception as e:
        print(f"Error populating template: {str(e)}")
        return False

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python populate_wbs_template.py <json_data_file> <output_path>")
        sys.exit(1)
        
    json_file = sys.argv[1]
    output_path = sys.argv[2]
    
    try:
        with open(json_file, 'r') as f:
            wbs_data = json.load(f)
        success = populate_wbs_template(wbs_data, output_path, output_path)
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1) 
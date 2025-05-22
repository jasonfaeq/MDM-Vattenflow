from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
import os
import tempfile
import json
from fastapi.responses import StreamingResponse
import io
import sys

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://mdm-vattenflow.vercel.app"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add region full name mapping at the top
REGION_FULL_NAMES = {
    'DE': 'Germany',
    'NL': 'Netherlands',
    'UK': 'Nordic',
    'SE': 'Nordic',
    'DK': 'Nordic'
}

def get_full_region_name(region_code):
    return REGION_FULL_NAMES.get(region_code, region_code)

def combine_code_label(code, label):
    if code and label and str(code) not in str(label):
        return f"{code} {label}"
    elif label:
        return label
    elif code:
        return code
    return ''

def get_template_path(region):
    """Get the appropriate template path based on region."""
    # For now, using the same template for all regions
    # In the future, you can add region-specific templates here
    return os.path.join(os.path.dirname(__file__), 'templates', 'wbs_template_actual.xlsm')

def fill_excel_template(template_path, output_path, wbs_data):
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb['Shared Template']
    start_row = 18
    ws['D7'] = wbs_data[0].get('requesterDisplayName', '') if wbs_data else ''
    ws['D8'] = wbs_data[0].get('responsiblePerson', '') if wbs_data else ''
    # Helper to write integer if possible
    def write_int(ws, row, col, value):
        try:
            if value is not None and value != '':
                ws.cell(row=row, column=col, value=int(value))
            else:
                ws.cell(row=row, column=col, value=value)
        except Exception:
            ws.cell(row=row, column=col, value=value)

    for i, element in enumerate(wbs_data):
        row = start_row + i
        # Use full region name
        region_code = element.get('region', '')
        ws.cell(row=row, column=1, value=get_full_region_name(region_code))
        ws.cell(row=row, column=2, value=element.get('type', ''))
        ws.cell(row=row, column=3, value=element.get('system', 'KIS'))
        # Combine code and label for controlling area
        controlling_area = combine_code_label(
            element.get('controllingArea', ''),
            element.get('controllingAreaLabel', '')
        )
        ws.cell(row=row, column=4, value=controlling_area)
        write_int(ws, row, 5, element.get('companyCode', ''))
        ws.cell(row=row, column=6, value=element.get('projectName', ''))
        ws.cell(row=row, column=7, value=element.get('projectDefinition', ''))
        write_int(ws, row, 8, element.get('level', ''))
        ws.cell(row=row, column=9, value=element.get('projectType', ''))
        write_int(ws, row, 10, element.get('investmentProfile', ''))
        write_int(ws, row, 11, element.get('responsibleProfitCenter', ''))
        write_int(ws, row, 12, element.get('responsibleCostCenter', ''))
        ws.cell(row=row, column=13, value='x' if element.get('planningElement') else '')
        ws.cell(row=row, column=14, value='x' if element.get('rubricElement') else '')
        ws.cell(row=row, column=15, value='x' if element.get('billingElement') else '')
        percent = element.get('settlementRulePercent', '')
        if percent not in ('', None):
            try:
                ws.cell(row=row, column=16, value=float(percent) / 100)
            except Exception:
                ws.cell(row=row, column=16, value=percent)
        else:
            ws.cell(row=row, column=16, value='')
        ws.cell(row=row, column=17, value=element.get('settlementRuleGoal', ''))
        ws.cell(row=row, column=18, value=element.get('projectProfile', ''))
        ws.cell(row=row, column=19, value=element.get('responsiblePerson', ''))
        ws.cell(row=row, column=20, value=element.get('userId', ''))
        write_int(ws, row, 21, element.get('employmentNumber', ''))
        # Combine code and label for functional area
        functional_area = combine_code_label(
            element.get('functionalArea', ''),
            element.get('functionalAreaLabel', '')
        )
        ws.cell(row=row, column=22, value=functional_area)
        ws.cell(row=row, column=24, value=element.get('comment', ''))
        ws.cell(row=row, column=25, value=element.get('tm1Project', ''))
        # Write tgPhase as integer if possible
        write_int(ws, row, 26, element.get('tgPhase', ''))
        # Project specification: if value is 'ASSET', write 'Asset' instead
        project_spec = element.get('projectSpec', '')
        if isinstance(project_spec, str) and project_spec.strip().upper() == 'ASSET':
            ws.cell(row=row, column=27, value='Asset')
        else:
            ws.cell(row=row, column=27, value=project_spec)
        ws.cell(row=row, column=28, value=element.get('motherCode', ''))
    wb.save(output_path)

@app.post("/export")
async def export_excel(payload: dict):
    wbs_data = payload.get("wbsData", [])
    if not wbs_data:
        raise HTTPException(status_code=400, detail="No WBS data provided")

    try:
        # Get the region from the first WBS element
        region = wbs_data[0].get('region')
        if not region:
            raise HTTPException(status_code=400, detail="Region not specified in WBS data")

        # Get the appropriate template for this region
        template_path = get_template_path(region)
        
        # Create a temporary file for the output
        with tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False) as tmp:
            # Fill the template with the WBS data
            fill_excel_template(template_path, tmp.name, wbs_data)
            tmp.seek(0)
            excel_data = tmp.read()
            
        # Clean up the temporary file
        os.unlink(tmp.name)
        
        # Return the filled template as a downloadable file
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            headers={"Content-Disposition": "attachment; filename=wbs_export.xlsm"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == '__main__':
    if len(sys.argv) != 4:
        print('Usage: python fill_excel_template.py <template_path> <output_path> <wbs_data_json>')
        sys.exit(1)
    template_path = sys.argv[1]
    output_path = sys.argv[2]
    wbs_data_json = sys.argv[3]
    with open(wbs_data_json, 'r', encoding='utf-8') as f:
        wbs_data = json.load(f)
    fill_excel_template(template_path, output_path, wbs_data)
    print('Excel file filled and saved successfully.') 